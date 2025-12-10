"""
ComprehensiveExcelGenerator: Formats and outputs all data into Excel workbook
"""
from typing import Dict, Any, List, Optional
from datetime import datetime, timedelta
from dateutil.parser import parse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import math


class ComprehensiveExcelGenerator:
    """Generates comprehensive Excel workbook with all forensic economic analysis data."""
    
    def __init__(self):
        self.workbook = None
        self.output_path = None
    
    def generate_excel_report(self, 
                             person_profile: Dict[str, Any],
                             life_expectancy_data: Dict[str, Any],
                             worklife_data: Dict[str, Any],
                             wage_data: Dict[str, Any],
                             discount_rate_data: Dict[str, Any],
                             timeline: List[Dict[str, Any]],
                             earnings_table: List[Dict[str, Any]],
                             pv_table: List[Dict[str, Any]],
                             output_path: str = None) -> str:
        """
        Generate comprehensive Excel report.
        
        Args:
            person_profile: Person data
            life_expectancy_data: Life expectancy calculations
            worklife_data: Work-life expectancy data
            wage_data: Wage growth data
            discount_rate_data: Discount rate data
            timeline: Age timeline
            earnings_table: Projected earnings table
            pv_table: Present value table
            output_path: Output file path (optional)
            
        Returns:
            Path to generated Excel file
        """
        # Create workbook
        self.workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Generate output filename if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            name = person_profile.get('name', 'Person').replace(' ', '_')
            output_path = f"forensic_economic_report_{name}_{timestamp}.xlsx"
        
        self.output_path = output_path
        
        # Create main detailed earnings loss sheet (matching screenshot format)
        self._create_detailed_earnings_loss_sheet(
            person_profile, worklife_data, wage_data, discount_rate_data, 
            timeline, earnings_table, pv_table
        )
        
        # Create additional worksheets
        self._create_personal_data_sheet(person_profile)
        self._create_life_expectancy_sheet(life_expectancy_data, worklife_data)
        self._create_wage_growth_sheet(wage_data)
        self._create_discount_rate_sheet(discount_rate_data)
        self._create_summary_sheet(person_profile, pv_table, life_expectancy_data, 
                                  worklife_data, wage_data, discount_rate_data)
        
        # Save workbook
        self.workbook.save(output_path)
        return output_path
    
    def _create_detailed_earnings_loss_sheet(self,
                                            person_profile: Dict[str, Any],
                                            worklife_data: Dict[str, Any],
                                            wage_data: Dict[str, Any],
                                            discount_rate_data: Dict[str, Any],
                                            timeline: List[Dict[str, Any]],
                                            earnings_table: List[Dict[str, Any]],
                                            pv_table: List[Dict[str, Any]]):
        """Create the main detailed earnings loss sheet matching the screenshot format."""
        ws = self.workbook.create_sheet("Earnings Loss", 0)  # First sheet
        
        # Parse dates
        dod = parse(person_profile['dod'])
        dob = parse(person_profile['dob'])
        age_at_death = person_profile.get('age_at_death', 0)
        base_salary = person_profile.get('annual_salary', 0)
        growth_rate = wage_data.get('annual_growth_rate', 0.025)
        discount_rate = discount_rate_data.get('discount_rate', 0.045)
        
        # Calculate portion of year for death date
        # Days from death date to end of year (including death day)
        days_in_year = 365 if dod.year % 4 != 0 else 366
        day_of_year = dod.timetuple().tm_yday
        days_remaining = days_in_year - day_of_year + 1
        portion_first_year = round(days_remaining / days_in_year, 2)
        
        # Calculate work-life end date
        worklife_years = worklife_data.get('worklife_expectancy', 0)
        
        # Calculate end date for work-life (from death date)
        worklife_days = int(worklife_years * 365.25)
        worklife_end_date = dod + timedelta(days=worklife_days)
        
        # Calculate portion of last year (from start of year to work-life end)
        days_in_last_year = 365 if worklife_end_date.year % 4 != 0 else 366
        day_of_year_end = worklife_end_date.timetuple().tm_yday
        portion_last_year = round(day_of_year_end / days_in_last_year, 2)
        portion_last_year = min(1.0, portion_last_year)  # Ensure it doesn't exceed 1.0
        
        # Header information
        row = 1
        ws.cell(row, 1, "Name:")
        ws.cell(row, 2, person_profile.get('name', ''))
        row += 1
        
        ws.cell(row, 1, "Item:")
        ws.cell(row, 2, "Earnings Loss (More Conservative)")
        row += 1
        
        ws.cell(row, 1, "Present Value Date:")
        ws.cell(row, 2, f"Month --> {dod.strftime('%b')}")
        ws.cell(row, 3, f"Day --> {dod.day}")
        ws.cell(row, 4, f"Year --> {dod.year}")
        
        # Format the present value date cells
        ws.cell(row, 2).alignment = Alignment(horizontal="left")
        ws.cell(row, 3).alignment = Alignment(horizontal="left")
        ws.cell(row, 4).alignment = Alignment(horizontal="left")
        row += 2
        
        # Key parameters - calculate final cumulative PV after we build the table
        # For now, we'll calculate it in the loop
        
        ws.cell(row, 1, "Base Value:")
        ws.cell(row, 2, f"${base_salary:,.2f}")
        row += 1
        
        ws.cell(row, 1, "Discount rate:")
        ws.cell(row, 2, f"{discount_rate * 100:.2f}%")
        row += 1
        
        ws.cell(row, 1, "Annual growth rate:")
        ws.cell(row, 2, f"{growth_rate * 100:.2f}%")
        row += 1
        
        # Add source for annual growth rate
        growth_source = wage_data.get('data_source', 'N/A')
        growth_method = wage_data.get('calculation_method', '')
        if growth_method:
            growth_source_text = f"{growth_source} ({growth_method})"
        else:
            growth_source_text = growth_source
        ws.cell(row, 1, "Annual growth rate source:")
        ws.cell(row, 2, growth_source_text)
        row += 1
        
        # Placeholder for cumulative PV - will update after calculation
        cumulative_pv_cell = ws.cell(row, 1, "Cumulative Present Value:")
        cumulative_pv_value_cell = ws.cell(row, 2, "$0")
        row += 2
        
        # Table headers
        headers = [
            "Age", "Start Date", "Year Number", "Portion of Year",
            "Full Year Value", "Actual Value", "Cumulative Value",
            "Discount Factor", "Present Value", "Cumulative Present Value"
        ]
        
        header_row = row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
        
        # Filter pv_table to only include work-life years
        worklife_pv_table = [entry for entry in pv_table if entry.get('projected_earnings', 0) > 0]
        
        # Calculate data for each year
        cumulative_value = 0
        cumulative_pv = 0
        accumulated_years = 0  # Track accumulated time for discount calculation
        
        for idx, entry in enumerate(worklife_pv_table):
            year_offset = entry['year_offset']
            age = entry['age']
            year = entry['year']
            full_year_value = entry['projected_earnings']
            
            # Calculate portion of year
            if idx == 0:
                # First year - partial year from death date to end of year
                portion_of_year = portion_first_year
            elif idx == len(worklife_pv_table) - 1:
                # Last year - partial year from start of year to work-life end date
                portion_of_year = portion_last_year
            else:
                # Full years
                portion_of_year = 1.00
            
            # Calculate actual value (pro-rated based on portion of year)
            actual_value = full_year_value * portion_of_year
            cumulative_value += actual_value
            
            # Calculate discount factor based on accumulated time from present value date
            # Discount at the midpoint of each payment period
            # First period: discount at midpoint of first portion (portion/2)
            # Subsequent periods: discount at accumulated time + midpoint of current portion
            
            # Calculate discount period (time from present value date to midpoint of payment)
            if idx == 0:
                # First period - discount at midpoint
                discount_period = accumulated_years + (portion_of_year / 2.0)
            else:
                # Subsequent periods - add accumulated time plus midpoint
                discount_period = accumulated_years + (portion_of_year / 2.0)
            
            # Calculate discount factor
            # If discount period is very small (< 0.01), treat as present value (factor = 1.0)
            if discount_period < 0.01:
                discount_factor = 1.0
            else:
                discount_factor = 1.0 / ((1 + discount_rate) ** discount_period)
            
            # Update accumulated years for next iteration (after calculating discount)
            accumulated_years += portion_of_year
            
            # Calculate present value
            present_value = actual_value * discount_factor
            cumulative_pv += present_value
            
            # Calculate age with decimal (age at start of period)
            # Age should reflect the actual age at the start of each period
            # For the first period, it's the age at death
            # For subsequent periods, age increases based on time elapsed
            if idx == 0:
                # First period starts at death date
                age_decimal = float(age_at_death)
            else:
                # Subsequent periods - age increases by the sum of previous portions
                # Calculate cumulative time elapsed
                previous_portions = portion_first_year if idx == 1 else (portion_first_year + (idx - 1))
                age_decimal = float(age_at_death) + previous_portions
            
            # Write row data
            ws.cell(row, 1, round(age_decimal, 1))  # Age
            ws.cell(row, 2, year)  # Start Date (year)
            ws.cell(row, 3, float(year_offset + 1))  # Year Number (starts at 1.0)
            ws.cell(row, 4, portion_of_year)  # Portion of Year
            ws.cell(row, 5, round(full_year_value, 2))  # Full Year Value
            ws.cell(row, 6, round(actual_value, 2))  # Actual Value
            ws.cell(row, 7, round(cumulative_value, 2))  # Cumulative Value
            ws.cell(row, 8, round(discount_factor, 5))  # Discount Factor
            ws.cell(row, 9, round(present_value, 2))  # Present Value
            ws.cell(row, 10, round(cumulative_pv, 2))  # Cumulative Present Value
            
            # Format cells
            ws.cell(row, 1).number_format = '0.0'  # Age
            ws.cell(row, 2).number_format = '0'  # Start Date
            ws.cell(row, 3).number_format = '0.0'  # Year Number
            ws.cell(row, 4).number_format = '0.00'  # Portion of Year
            ws.cell(row, 5).number_format = '$#,##0.00'  # Full Year Value
            ws.cell(row, 6).number_format = '$#,##0.00'  # Actual Value
            ws.cell(row, 7).number_format = '$#,##0.00'  # Cumulative Value
            ws.cell(row, 8).number_format = '0.00000'  # Discount Factor
            ws.cell(row, 9).number_format = '$#,##0.00'  # Present Value
            ws.cell(row, 10).number_format = '$#,##0.00'  # Cumulative Present Value
            
            # Add borders
            for col in range(1, 11):
                ws.cell(row, col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            row += 1
        
        # Update cumulative present value in header
        cumulative_pv_value_cell.value = f"${cumulative_pv:,.0f}"
        
        # Set column widths
        ws.column_dimensions['A'].width = 10  # Age
        ws.column_dimensions['B'].width = 12  # Start Date
        ws.column_dimensions['C'].width = 12  # Year Number
        ws.column_dimensions['D'].width = 15  # Portion of Year
        ws.column_dimensions['E'].width = 15  # Full Year Value
        ws.column_dimensions['F'].width = 15  # Actual Value
        ws.column_dimensions['G'].width = 18  # Cumulative Value
        ws.column_dimensions['H'].width = 15  # Discount Factor
        ws.column_dimensions['I'].width = 15  # Present Value
        ws.column_dimensions['J'].width = 22  # Cumulative Present Value
    
    def _create_personal_data_sheet(self, person_profile: Dict[str, Any]):
        """Create personal data worksheet."""
        ws = self.workbook.create_sheet("Personal Data")
        
        # Header
        ws['A1'] = "Personal Information"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        # Data
        row = 3
        fields = [
            ('Name', 'name'),
            ('Date of Birth', 'dob'),
            ('Date of Death', 'dod'),
            ('Age at Death', 'age_at_death'),
            ('Occupation', 'occupation'),
            ('Annual Salary', 'annual_salary'),
            ('Sex', 'sex'),
            ('Education Level', 'education_level'),
            ('Home County', 'home_county'),
            ('Home State', 'home_state'),
            ('Status', 'status')
        ]
        
        for label, key in fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            value = person_profile.get(key, 'N/A')
            if key == 'annual_salary' and isinstance(value, (int, float)):
                ws[f'B{row}'] = f"${value:,.2f}"
            else:
                ws[f'B{row}'] = str(value)
            row += 1
        
        # Format columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_life_expectancy_sheet(self, life_expectancy_data: Dict[str, Any],
                                     worklife_data: Dict[str, Any]):
        """Create life expectancy and work-life expectancy worksheet."""
        ws = self.workbook.create_sheet("Life & Work-Life Expectancy")
        
        # Header
        ws['A1'] = "Life Expectancy Analysis"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        # Life Expectancy
        row = 3
        ws[f'A{row}'] = "Life Expectancy"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        life_fields = [
            ('Age at Death', 'age_at_death'),
            ('Remaining Life Expectancy (years)', 'remaining_life_expectancy'),
            ('Total Expected Lifespan', 'total_expected_lifespan'),
            ('Sex', 'sex'),
            ('Data Source', 'data_source')
        ]
        
        for label, key in life_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = str(life_expectancy_data.get(key, 'N/A'))
            row += 1
        
        row += 1
        
        # Work-Life Expectancy
        ws[f'A{row}'] = "Work-Life Expectancy"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        worklife_fields = [
            ('Age at Death', 'age_at_death'),
            ('Work-Life Expectancy (years)', 'worklife_expectancy'),
            ('Education Level', 'education_level'),
            ('Data Source', 'data_source')
        ]
        
        for label, key in worklife_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = str(worklife_data.get(key, 'N/A'))
            row += 1
        
        # Format columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 30
    
    def _create_wage_growth_sheet(self, wage_data: Dict[str, Any]):
        """Create wage growth data worksheet."""
        ws = self.workbook.create_sheet("Wage Growth")
        
        # Header
        ws['A1'] = "Salary Growth Rate Analysis"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        row = 3
        fields = [
            ('Occupation', 'occupation'),
            ('County', 'county'),
            ('State', 'state'),
            ('Annual Growth Rate', 'annual_growth_rate'),
            ('Annual Growth Percent', 'annual_growth_percent'),
            ('Data Source', 'data_source'),
            ('Calculation Method', 'calculation_method')
        ]
        
        for label, key in fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            value = wage_data.get(key, 'N/A')
            if key == 'annual_growth_rate':
                ws[f'B{row}'] = f"{value:.4f}"
            elif key == 'annual_growth_percent':
                ws[f'B{row}'] = f"{value}%"
            else:
                ws[f'B{row}'] = str(value)
            row += 1
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    def _create_discount_rate_sheet(self, discount_rate_data: Dict[str, Any]):
        """Create discount rate worksheet."""
        ws = self.workbook.create_sheet("Discount Rate")
        
        # Header
        ws['A1'] = "Discount Rate Analysis"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        row = 3
        fields = [
            ('Discount Rate', 'discount_rate'),
            ('Discount Rate Percent', 'discount_rate_percent'),
            ('Rate Type', 'rate_type'),
            ('Data Source', 'data_source'),
            ('Fetch Status', 'fetch_status')
        ]
        
        for label, key in fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            value = discount_rate_data.get(key, 'N/A')
            if key == 'discount_rate':
                ws[f'B{row}'] = f"{value:.4f}"
            elif key == 'discount_rate_percent':
                ws[f'B{row}'] = f"{value}%"
            else:
                ws[f'B{row}'] = str(value)
            row += 1
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    def _create_summary_sheet(self, person_profile: Dict[str, Any],
                             pv_table: List[Dict[str, Any]],
                             life_expectancy_data: Dict[str, Any],
                             worklife_data: Dict[str, Any],
                             wage_data: Dict[str, Any],
                             discount_rate_data: Dict[str, Any]):
        """Create summary worksheet with key results."""
        ws = self.workbook.create_sheet("Summary", 0)  # Insert at beginning
        
        # Title
        ws['A1'] = "Forensic Economic Loss Analysis - Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        # Report date
        ws['A2'] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:B2')
        
        row = 4
        
        # Total Economic Loss
        total_loss = pv_table[-1]['cumulative_present_value'] if pv_table else 0
        ws[f'A{row}'] = "TOTAL ECONOMIC LOSS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'B{row}'] = f"${total_loss:,.2f}"
        ws[f'B{row}'].font = Font(bold=True, size=14, color="C00000")
        row += 2
        
        # Key Metrics
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        metrics = [
            ('Name', person_profile.get('name', 'N/A')),
            ('Age at Death', f"{person_profile.get('age_at_death', 'N/A')} years"),
            ('Remaining Life Expectancy', f"{life_expectancy_data.get('remaining_life_expectancy', 0):.2f} years"),
            ('Work-Life Expectancy', f"{worklife_data.get('worklife_expectancy', 0):.2f} years"),
            ('Base Annual Salary', f"${person_profile.get('annual_salary', 0):,.2f}"),
            ('Annual Salary Growth Rate', f"{wage_data.get('annual_growth_percent', 0)}%"),
            ('Discount Rate', f"{discount_rate_data.get('discount_rate_percent', 0)}%"),
            ('Years of Projected Earnings', f"{len([e for e in pv_table if e['projected_earnings'] > 0])} years")
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = str(value)
            row += 1
        
        # Format columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
